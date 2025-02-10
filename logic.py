import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import logging
from rapidfuzz import fuzz

# Konfiguroidaan lokitus, jotta näemme mitä koodissa tapahtuu
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class ExcelProcessor:
    def __init__(self):
        # Alustetaan viite- ja tarjousten avainsarakkeet
        self.ref_key_column = None
        self.offer_key_column = None
        # Lista sarakkeista, jotka halutaan ottaa mukaan yhdistämisessä
        self.selected_ref_columns = []

    def process_files(self, reference_file, offer_file, reference_column, competitor_column):
        """
        Päämetodi, joka suorittaa tiedostojen prosessoinnin ja yhdistämisen.
        """
        self.ref_key_column = reference_column
        self.offer_key_column = competitor_column

        # Varmistetaan, ettei viiteavainsarake ole mukana käyttäjän valituissa sarakkeissa
        if self.ref_key_column in self.selected_ref_columns:
            self.selected_ref_columns.remove(self.ref_key_column)
            logging.info(f"Removed reference key column '{self.ref_key_column}' from selected columns.")

        # Ladataan ja valmistellaan tiedostot
        df_reference, df_offer = self.load_and_prepare_files(reference_file, offer_file)
        # Suoritetaan tiedostojen yhdistäminen
        merged_df = self.merge_data(df_reference, df_offer)
        # Tallennetaan yhdistetty data uuteen Excel-tiedostoon
        output_path, missing_count = self.save_to_excel(offer_file, merged_df)
        logging.info(f"Processing complete. Output saved to '{output_path}'. Missing count: {missing_count}")
        return output_path, missing_count

    def load_and_prepare_files(self, reference_file, offer_file):
        """
        Lataa Excel-tiedostot Pandas DataFrameihin ja tarkistaa, että tarvittavat sarakkeet ovat olemassa.
        """
        try:
            df_reference = pd.read_excel(reference_file, dtype=str)
            logging.info(f"Reference file '{reference_file}' loaded successfully.")
        except Exception as e:
            logging.error(f"Could not read the reference file: {e}")
            raise ValueError(f"Could not read the reference file: {e}")

        try:
            df_offer = pd.read_excel(offer_file, dtype=str)
            logging.info(f"Offer file '{offer_file}' loaded successfully.")
        except Exception as e:
            logging.error(f"Could not read the offer file: {e}")
            raise ValueError(f"Could not read the offer file: {e}")

        # Tarkistetaan, että viiteavaimesarake löytyy viitetiedostosta
        if self.ref_key_column not in df_reference.columns:
            logging.error(f"Chosen reference key '{self.ref_key_column}' not found in reference file.")
            raise ValueError(f"Chosen reference key '{self.ref_key_column}' not found in reference file.")
        # Tarkistetaan, että tarjousavaimesarake löytyy tarjoustiedostosta
        if self.offer_key_column not in df_offer.columns:
            logging.error(f"Chosen offer key '{self.offer_key_column}' not found in offer file.")
            raise ValueError(f"Chosen offer key '{self.offer_key_column}' not found in offer file.")

        # Tarkistetaan, että käyttäjän valitsemat sarakkeet löytyvät viitetiedostosta
        for col in self.selected_ref_columns:
            if col not in df_reference.columns:
                logging.error(f"Selected column '{col}' not in reference file.")
                raise ValueError(f"Selected column '{col}' not in reference file.")
        
        return df_reference, df_offer

    def merge_data(self, df_reference, df_offer):
        """
        Yhdistää viite- ja tarjoustiedot useilla eri strategioilla.
        """
        # 1) Poistetaan duplikaatit viitetiedostosta viiteavaimen perusteella
        df_reference = df_reference.drop_duplicates(subset=[self.ref_key_column], keep="first").copy()
        logging.info(f"Deduplicated reference data based on '{self.ref_key_column}'.")

        # 2) Poistetaan välilyönnit ja trimmaillaan tarjousavaimen arvot
        df_offer[self.offer_key_column] = df_offer[self.offer_key_column].str.replace(" ", "").str.strip()
        logging.info(f"Removed spaces and stripped offer key column '{self.offer_key_column}'.")

        # 3) Säilytetään tarjoustiedoston alkuperäinen rivijärjestys
        df_offer["_original_order"] = range(len(df_offer))

        # Määritellään yhdistettävät sarakkeet; jos käyttäjä on valinnut lisäsarakkeita, otetaan ne mukaan
        if self.selected_ref_columns:
            columns_to_merge = [self.ref_key_column] + self.selected_ref_columns
        else:
            columns_to_merge = [self.ref_key_column]

        # 4) Uudelleennimetään viitetiedoston sarakkeet, jos tarjoustiedostossa on samannimisiä sarakkeita
        rename_dict = {col: f"{col} (referenssi)" for col in columns_to_merge if col in df_offer.columns}
        df_reference.rename(columns=rename_dict, inplace=True)
        logging.info(f"Renamed columns in reference data: {rename_dict}")
        columns_to_merge = [rename_dict.get(col, col) for col in columns_to_merge]

        # 5) Suoritetaan ensimmäinen yhdistys Pandas merge-toiminnolla
        merged_df = pd.merge(
            df_offer,
            df_reference[columns_to_merge],
            left_on=self.offer_key_column,
            right_on=columns_to_merge[0],
            how="left",
            sort=False,
            indicator=True  # Lisää '_merge' sarakkeen, joka kertoo yhdistämisen tuloksen
        )
        logging.info("Performed initial merge.")

        # Alustetaan 'used_code'-sarake, joka säilyttää alkuperäisen tarjousavaimen
        merged_df['used_code'] = merged_df[self.offer_key_column]

        # 6) Ensimmäinen vaihtoehtoinen strategia: yritetään yhdistää lisäämällä tarjousavaimeen eteen '0'
        unmatched = merged_df['_merge'] == 'left_only'
        if unmatched.any():
            logging.info(f"Found {unmatched.sum()} unmatched records. Attempting match with a leading '0'.")
            df_unmatched = df_offer.loc[unmatched].copy()
            df_unmatched[self.offer_key_column] = '0' + df_unmatched[self.offer_key_column].astype(str)
            merged_unmatched = pd.merge(
                df_unmatched,
                df_reference[columns_to_merge],
                left_on=self.offer_key_column,
                right_on=columns_to_merge[0],
                how="left",
                sort=False,
                indicator=True
            )
            logging.info("Performed secondary merge with leading '0'.")
            unmatched_indices = merged_df[unmatched].index
            for idx, (_, row) in zip(unmatched_indices, merged_unmatched.iterrows()):
                if row['_merge'] == 'both':
                    # Päivitetään yhdistetyn DataFrame:n tiedot, jos löytyy osuma
                    for col in columns_to_merge:
                        merged_df.at[idx, col] = row[col]
                    merged_df.at[idx, '_merge'] = 'both'
                    merged_df.at[idx, 'used_code'] = row[self.offer_key_column]

        # 7) Toinen vaihtoehtoinen strategia: yritetään etuliitteen mukaista vertailua
        still_unmatched = merged_df['_merge'] == 'left_only'
        if still_unmatched.any():
            logging.info(f"{still_unmatched.sum()} records still unmatched. Trying alternative prefix matching.")
            for idx in merged_df[still_unmatched].index:
                offer_code = merged_df.at[idx, self.offer_key_column]
                alt_match = self.find_alternative_match(offer_code, df_reference[self.ref_key_column])
                if alt_match is not None:
                    # Etsitään viiterivi, jossa siivotut arvot vastaavat toisiaan
                    ref_row = df_reference[
                        df_reference[self.ref_key_column].str.replace(" ", "").str.strip().str.lower() == alt_match
                    ]
                    if not ref_row.empty:
                        for col in columns_to_merge:
                            merged_df.at[idx, col] = ref_row.iloc[0][col]
                        merged_df.at[idx, '_merge'] = 'both'
                        merged_df.at[idx, 'used_code'] = alt_match

        # 8) Kolmas vaihtoehtoinen strategia: käytetään fuzzy matching -menetelmää
        still_unmatched = merged_df['_merge'] == 'left_only'
        if still_unmatched.any():
            logging.info(f"{still_unmatched.sum()} records still unmatched. Trying fuzzy matching.")
            for idx in merged_df[still_unmatched].index:
                offer_code = merged_df.at[idx, self.offer_key_column]
                alt_match = self.find_fuzzy_match(offer_code, df_reference[self.ref_key_column])
                if alt_match is not None:
                    ref_row = df_reference[
                        df_reference[self.ref_key_column].str.replace(" ", "").str.strip().str.lower() == alt_match
                    ]
                    if not ref_row.empty:
                        for col in columns_to_merge:
                            merged_df.at[idx, col] = ref_row.iloc[0][col]
                        merged_df.at[idx, '_merge'] = 'both'
                        merged_df.at[idx, 'used_code'] = alt_match

        # 9) Lisätään 'matched'-sarake, joka kertoo onko rivi yhdistetty, palautetaan alkuperäinen rivijärjestys
        merged_df['matched'] = merged_df['_merge'] == 'both'
        merged_df.sort_values("_original_order", inplace=True)
        merged_df.drop(columns=["_merge", "_original_order"], inplace=True)
        logging.info("Restored original row order and removed helper columns.")

        return merged_df

    def find_alternative_match(self, offer_code, ref_series):
        """
        Yrittää löytää vaihtoehtoisen matchin, jossa tarkastellaan alkiota, 
        joka on toinen koodin alkuosa. Palauttaa siivotun version referenssikoodista, jos löytyy.
        """
        cleaned_offer = str(offer_code).replace(" ", "").strip().lower()
        for ref_code in ref_series.dropna().unique():
            cleaned_ref = str(ref_code).replace(" ", "").strip().lower()
            if cleaned_offer.startswith(cleaned_ref) or cleaned_ref.startswith(cleaned_offer):
                return cleaned_ref
        return None

    def find_fuzzy_match(self, offer_code, ref_series, threshold=80):
        """
        Käyttää fuzzy matching -menetelmää (rapidfuzz) etsimään paras mahdollinen osuma.
        Palauttaa matchatun koodin, jos pistemäärä ylittää asetetun kynnyksen.
        """
        cleaned_offer = str(offer_code).replace(" ", "").strip().lower()
        best_match = None
        best_score = 0
        for ref_code in ref_series.dropna().unique():
            cleaned_ref = str(ref_code).replace(" ", "").strip().lower()
            score = fuzz.token_sort_ratio(cleaned_offer, cleaned_ref)
            if score > best_score:
                best_score = score
                best_match = cleaned_ref
        if best_score >= threshold:
            return best_match
        return None

    def save_to_excel(self, offer_file, merged_df):
        """
        Tallentaa yhdistetyn DataFrame:n takaisin Excel-tiedostoon.
        Lisää uudet sarakkeet, tyylittelee ne ja tallentaa tiedoston aikaleimalla.
        """
        # Ladataan alkuperäinen workbook openpyxl:lla
        wb = load_workbook(offer_file)
        ws = wb.active
        logging.info(f"Loaded workbook '{offer_file}' for saving.")

        # Luetaan alkuperäiset sarakkeet tarjoustiedostosta
        original_offer_df = pd.read_excel(offer_file, dtype=str)
        original_cols = set(original_offer_df.columns)

        # Määritellään lisättävät sarakkeet: käyttäjän valitsemat referenssisarakkeet, jos niitä ei ole alkuperäisessä tiedostossa
        new_columns = []
        for col in self.selected_ref_columns:
            if col not in original_cols:
                new_columns.append(col)
        # Lisätään myös 'used_code'-sarake
        new_columns.append('used_code')

        # Lasketaan uusi sarakealku, josta uudet sarakkeet lisätään
        start_col = ws.max_column + 1
        logging.info(f"Adding new columns starting at column {start_col}.")

        matched_list = merged_df['matched'].tolist()

        # Lisätään uudet sarakkeet ja täytetään niillä dataa
        self.add_new_columns(ws, merged_df, new_columns, start_col)
        # Muotoillaan uudet sarakkeet (värit, reunat)
        self.style_new_columns(ws, start_col, len(new_columns), matched_list)
        # Asetetaan yhtenäinen sarakeleveys
        self.set_uniform_column_width(ws, 25)
        # Poistetaan mahdollinen jäädytetty paneeli
        ws.freeze_panes = None
        # Muotoillaan header-rivi (otsikot)
        self.style_header_row(ws)

        # Tallennetaan tiedosto uudella nimellä, joka sisältää aikaleiman
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        output_path = Path(offer_file).parent / f"MATCHED_{timestamp}.xlsx"
        wb.save(output_path)
        logging.info(f"Saved merged workbook to '{output_path}'.")

        # Lasketaan ja logitetaan yhdistämättömien rivien määrä
        matched_count = merged_df['matched'].sum()  # Osumien määrä
        total_rows = len(merged_df)
        unmatched_count = total_rows - matched_count
        logging.info(f"Count of missing matches: {unmatched_count}")

        return output_path, unmatched_count

    def add_new_columns(self, worksheet, merged_df, new_columns, start_col):
        """
        Lisää uudet sarakkeet Excel-työkirjaan ja täyttää ne datalla.
        Jos riviä ei ole yhdistetty, asetetaan soluun teksti "Ei vastaavaa".
        """
        # Kirjoitetaan sarakeotsikot
        for i, col_name in enumerate(new_columns, start=start_col):
            header_cell = worksheet.cell(row=1, column=i)
            header_cell.value = col_name
            logging.debug(f"Added header '{col_name}' at column {i}.")

        # Täytetään uudet sarakkeet datalla rivittäin
        for row_idx in range(len(merged_df)):
            excel_row = row_idx + 2  # Excelissä ensimmäinen rivi on header
            matched = merged_df['matched'].iloc[row_idx]
            for j, col_name in enumerate(new_columns, start=start_col):
                cell = worksheet.cell(row=excel_row, column=j)
                if col_name == 'used_code':
                    cell.value = merged_df['used_code'].iloc[row_idx]
                else:
                    if not matched:
                        cell.value = "Ei vastaavaa"
                    else:
                        value = merged_df[col_name].iloc[row_idx]
                        cell.value = value if pd.notna(value) else ""
                # Asetetaan solun tasoitus vasemmalle ja keskitetty vertikaalisesti
                cell.alignment = Alignment(horizontal="left", vertical="center")
        logging.info("Filled new columns with data, setting 'Ei vastaavaa' where applicable.")

    def style_new_columns(self, worksheet, start_col, num_cols, matched_list):
        """
        Muotoilee uudet sarakkeet: asettaa täyttövärit ja reunukset.
        Soluissa, joissa osumaa ei ole, käytetään punaista väriä; muuten vihreää.
        """
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        thin_side = Side(style="thin")
        thick_side = Side(style="thick")

        for col_idx in range(start_col, start_col + num_cols):
            # Tarkistetaan, onko kyseessä ensimmäinen tai viimeinen uusi sarake
            is_first_col = (col_idx == start_col)
            is_last_col = (col_idx == start_col + num_cols - 1)
            col_name = worksheet.cell(row=1, column=col_idx).value

            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if col_name == 'used_code':
                    cell.fill = green_fill if matched_list[row_idx - 2] else red_fill
                else:
                    cell.fill = red_fill if cell.value == "Ei vastaavaa" else green_fill

                # Asetetaan reunat: paksu reuna ensimmäisessä ja viimeisessä sarakkeessa
                left_border = thick_side if is_first_col else thin_side
                right_border = thick_side if is_last_col else thin_side
                cell.border = Border(
                    left=left_border,
                    right=right_border,
                    top=thin_side,
                    bottom=thin_side
                )
        logging.info("Styled new columns with fills and borders.")

    def set_uniform_column_width(self, worksheet, width=25):
        """
        Asettaa kaikkien sarakkeiden leveydeksi määritetyn arvon (oletus 25).
        """
        for col_idx in range(1, worksheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = width
        logging.info(f"Set all column widths to {width}.")

    def style_header_row(self, worksheet):
        """
        Muotoilee ensimmäisen rivin (otsikkorivin): tekee fontista lihavoidun ja keskittää tekstin.
        """
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = center_align
        logging.info("Styled header row with bold font and centered alignment.")

    def beautify_worksheet(self, ws):
        """
        Paikallinen metodi, johon voidaan myöhemmin lisätä lisää tyylittelyä ja
        parannuksia Excel-työkirjaan.
        """
        pass
